"""
visualizer.py — Owner Progression Graph
========================================
Generates a line chart showing cumulative point trajectories
for all owners from Match 17 onwards (post-bulk phase).
"""

import os
import openpyxl

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as ticker
except ImportError:
    print("❌ matplotlib not available.")
    raise

from config import LEDGER_FILE, OWNER_SHEET_PREFIX, BULK_CUTOFF_MATCH
from scoring_engine import load_owner_map


OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "owner_progression.png")

OWNER_COLORS = {
    "Kaushik":  "#E63946",
    "Vignesh":  "#E9C46A",
    "Aravinth": "#2A9D8F",
    "Sid":      "#F4A261",
    "Barath":   "#457B9D",
    "JR":       "#264653",
    "Venky":    "#A855F7",
}

MARKERS = ["o", "s", "^", "D", "v", "P", "X"]


def generate_progression():
    """Generate the owner progression chart."""
    try:
        wb = openpyxl.load_workbook(LEDGER_FILE, data_only=True)
    except FileNotFoundError:
        print("❌ Ledger file not found. Run --compute first.")
        return

    owner_map = load_owner_map()
    owners = list(owner_map.keys())

    # ── Collect match columns and per-owner points ──
    # We read from each owner sheet
    owner_data = {}  # {owner: [(match_label, cumulative_pts)]}

    for owner in owners:
        sheet_name = f"{OWNER_SHEET_PREFIX}{owner}"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Find match columns (starting after Pre-M16)
        match_cols = []
        for col in range(4, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).startswith("M"):
                match_cols.append((col, str(header)))
            elif header == "Pre-M16 Points":
                match_cols.append((col, "Pre-M16"))

        # Sum points per match column across all players
        cumulative = 0
        data_points = []
        for col, label in match_cols:
            match_total = 0
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if isinstance(val, (int, float)):
                    match_total += val
            cumulative += match_total
            data_points.append((label, match_total, cumulative))

        owner_data[owner] = data_points

    if not owner_data:
        print("❌ No data to plot.")
        return

    # ── Get all unique match labels in order ──
    all_labels = []
    seen = set()
    for owner in owners:
        for label, _, _ in owner_data.get(owner, []):
            if label not in seen:
                all_labels.append(label)
                seen.add(label)

    if not all_labels:
        print("❌ No match columns found in owner sheets.")
        return

    x = list(range(len(all_labels)))

    # ── Plot ──
    fig, (ax1, ax2) = plt.subplots(
        2, 1, figsize=(max(14, len(all_labels) * 0.8), 12),
        gridspec_kw={"height_ratios": [3, 1.2]},
        facecolor="#1A1A2E",
    )

    for ax in [ax1, ax2]:
        ax.set_facecolor("#16213E")
        ax.tick_params(colors="#E0E0E0", labelsize=9)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color("#444")
        ax.spines["bottom"].set_color("#444")
        ax.grid(True, alpha=0.15, color="white")

    # ── Top: Cumulative ──
    for idx, owner in enumerate(owners):
        color = OWNER_COLORS.get(owner, f"C{idx}")
        marker = MARKERS[idx % len(MARKERS)]
        data_points = owner_data.get(owner, [])
        label_to_cum = {dp[0]: dp[2] for dp in data_points}
        y_vals = [label_to_cum.get(lbl, None) for lbl in all_labels]

        # Forward-fill None values
        filled = []
        last_val = 0
        for v in y_vals:
            if v is not None:
                last_val = v
            filled.append(last_val)

        ax1.plot(x, filled, color=color, marker=marker, markersize=4,
                 linewidth=2.2, label=owner, markeredgecolor="white",
                 markeredgewidth=0.5, alpha=0.95)

        # Annotate final point
        ax1.annotate(f"{filled[-1]:.0f}", xy=(x[-1], filled[-1]),
                     xytext=(8, 0), textcoords="offset points",
                     fontsize=10, fontweight="bold", color=color, va="center")

    ax1.set_ylabel("Cumulative Points", fontsize=12, color="#E0E0E0", fontweight="bold")
    ax1.set_title("IPL Fantasy — Owner Points Progression",
                   fontsize=16, color="white", fontweight="bold", pad=15)
    ax1.set_xticks(x)
    ax1.set_xticklabels(all_labels, rotation=45, ha="right", fontsize=8)
    ax1.legend(loc="upper left", fontsize=10, facecolor="#16213E",
               edgecolor="#444", labelcolor="#E0E0E0", framealpha=0.9)
    ax1.yaxis.set_major_formatter(ticker.FuncFormatter(lambda v, _: f"{int(v):,}"))

    # ── Bottom: Per-match bars ──
    bar_width = 0.8 / len(owners)
    offsets = [(i - len(owners) / 2 + 0.5) * bar_width for i in range(len(owners))]

    for idx, owner in enumerate(owners):
        color = OWNER_COLORS.get(owner, f"C{idx}")
        data_points = owner_data.get(owner, [])
        label_to_match = {dp[0]: dp[1] for dp in data_points}
        y_vals = [label_to_match.get(lbl, 0) for lbl in all_labels]

        positions = [xi + offsets[idx] for xi in x]
        ax2.bar(positions, y_vals, bar_width, color=color, alpha=0.85,
                label=owner, edgecolor="none")

    ax2.set_xlabel("Match", fontsize=12, color="#E0E0E0", fontweight="bold")
    ax2.set_ylabel("Match Points", fontsize=10, color="#E0E0E0", fontweight="bold")
    ax2.set_title("Points Per Match", fontsize=12, color="#CCCCCC", pad=8)
    ax2.set_xticks(x)
    ax2.set_xticklabels(all_labels, rotation=45, ha="right", fontsize=8)
    ax2.legend(loc="upper left", fontsize=8, facecolor="#16213E",
               edgecolor="#444", labelcolor="#E0E0E0", framealpha=0.9,
               ncol=len(owners))

    plt.tight_layout(pad=2)
    plt.savefig(OUTPUT_FILE, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()

    print(f"✅ Progression chart saved to: {os.path.abspath(OUTPUT_FILE)}")


if __name__ == "__main__":
    generate_progression()
