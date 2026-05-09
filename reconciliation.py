"""
reconciliation.py — Data Integrity Verifier
============================================
Cross-checks the Excel ledger for consistency:
  1. Sum of each owner's player points == Owner total in leaderboard
  2. Sum of match-by-match entries == Player cumulative total
  3. Detects data drift from manual edits
"""

import openpyxl
from config import LEDGER_FILE, SHEET_LEADERBOARD, SHEET_GLOBAL_LEDGER, OWNER_SHEET_PREFIX
from scoring_engine import load_owner_map, build_player_to_owner


def run_tests() -> bool:
    """Run all reconciliation checks. Returns True if all pass."""
    try:
        wb = openpyxl.load_workbook(LEDGER_FILE, data_only=True)
    except FileNotFoundError:
        print("❌ Ledger file not found. Run --compute first.")
        return False

    owner_map = load_owner_map()
    all_passed = True
    checks_run = 0
    checks_passed = 0

    print("=" * 60)
    print("🔍 IPL FANTASY — DATA RECONCILIATION")
    print("=" * 60)

    # ── Check 1: Owner totals from player sheets match leaderboard ──
    print("\n📊 Check 1: Owner sheet totals vs Leaderboard")
    ws_lb = wb[SHEET_LEADERBOARD]
    leaderboard = {}
    for row in range(2, ws_lb.max_row + 1):
        owner = ws_lb.cell(row=row, column=2).value
        total = ws_lb.cell(row=row, column=3).value
        if owner:
            leaderboard[owner] = total or 0

    for owner, players in owner_map.items():
        checks_run += 1
        sheet_name = f"{OWNER_SHEET_PREFIX}{owner}"
        if sheet_name not in wb.sheetnames:
            print(f"  ❌ {owner}: Sheet '{sheet_name}' missing!")
            all_passed = False
            continue

        ws = wb[sheet_name]
        sheet_total = 0
        for row in range(2, ws.max_row + 1):
            for col in range(4, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if isinstance(val, (int, float)):
                    sheet_total += val

        lb_total = leaderboard.get(owner, 0)
        if abs(sheet_total - lb_total) < 0.01:
            print(f"  ✅ {owner}: Sheet={sheet_total:.1f}, Leaderboard={lb_total:.1f}")
            checks_passed += 1
        else:
            print(f"  ❌ {owner}: Sheet={sheet_total:.1f} ≠ Leaderboard={lb_total:.1f} "
                  f"(diff={sheet_total - lb_total:.1f})")
            all_passed = False

    # ── Check 2: Global ledger per-player totals vs owner sheets ──
    print("\n📊 Check 2: Global ledger vs Owner sheets (per player)")
    ws_gl = wb[SHEET_GLOBAL_LEDGER]
    gl_player_totals = {}
    for row in range(2, ws_gl.max_row + 1):
        name = ws_gl.cell(row=row, column=3).value
        pts = ws_gl.cell(row=row, column=7).value
        if name and isinstance(pts, (int, float)):
            gl_player_totals[name] = gl_player_totals.get(name, 0) + pts

    player_to_owner = build_player_to_owner(owner_map)
    mismatches = 0
    for player_name, gl_total in gl_player_totals.items():
        checks_run += 1
        owner = player_to_owner.get(player_name)
        if not owner:
            continue

        sheet_name = f"{OWNER_SHEET_PREFIX}{owner}"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        sheet_total = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == player_name:
                for col in range(4, ws.max_column + 1):
                    val = ws.cell(row=row, column=col).value
                    if isinstance(val, (int, float)):
                        sheet_total += val
                break

        if abs(gl_total - sheet_total) < 0.01:
            checks_passed += 1
        else:
            print(f"  ❌ {player_name} ({owner}): Global={gl_total:.1f} ≠ Sheet={sheet_total:.1f}")
            mismatches += 1
            all_passed = False

    if mismatches == 0:
        print(f"  ✅ All {len(gl_player_totals)} players consistent")

    # ── Check 3: Draft integrity ──
    print("\n📊 Check 3: Draft integrity (no duplicate players)")
    checks_run += 1
    seen = {}
    dup_found = False
    for owner, players in owner_map.items():
        for p in players:
            if p in seen:
                print(f"  ❌ '{p}' is in both '{seen[p]}' and '{owner}'!")
                dup_found = True
                all_passed = False
            seen[p] = owner

    if not dup_found:
        print(f"  ✅ {len(seen)} unique players across {len(owner_map)} owners")
        checks_passed += 1

    # ── Summary ──
    print(f"\n{'=' * 60}")
    if all_passed:
        print(f"✅ ALL CHECKS PASSED ({checks_passed}/{checks_run})")
    else:
        print(f"❌ SOME CHECKS FAILED ({checks_passed}/{checks_run} passed)")
    print("=" * 60)

    return all_passed


if __name__ == "__main__":
    run_tests()
