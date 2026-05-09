"""
export_web_data.py — Exports all scoring data as JSON for the FE app
=====================================================================
Reads the ledger, owner map, replacement draft, milestone snapshots
and writes them into web/data/*.json for the static FE to consume.
"""

import json
import os
import shutil
from datetime import datetime

import openpyxl

from config import (
    LEDGER_FILE, OWNER_SHEET_PREFIX,
    SHEET_LEADERBOARD, SHEET_GLOBAL_LEDGER, SHEET_FIXTURES,
)
from scoring_engine import load_owner_map, build_player_to_owner
from replacement_manager import (
    load_replacement_draft, load_milestone_snapshot,
    apply_replacements_to_owner_map, compute_milestone_adjusted_points,
    compute_owner_totals_with_milestones, get_milestones,
)
from config import OWNER_MAP_JSON, REPLACEMENT_DRAFT_JSON, MILESTONE_SNAPSHOTS_DIR

WEB_DATA_DIR = os.path.join(os.path.dirname(__file__), "web", "data")
API_SNAPSHOT = os.path.join(os.path.dirname(__file__), "api_logs", "latest_snapshot.json")


def _write(filename, data):
    os.makedirs(WEB_DATA_DIR, exist_ok=True)
    path = os.path.join(WEB_DATA_DIR, filename)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"   📄 {filename}")


def export_meta():
    """Export snapshot timestamp and last update info."""
    meta = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "milestones": get_milestones(),
        "has_snapshot": os.path.exists(API_SNAPSHOT),
    }
    if os.path.exists(API_SNAPSHOT):
        with open(API_SNAPSHOT) as f:
            snap = json.load(f)
        meta["snapshot_timestamp"] = snap.get("timestamp", "unknown")
        meta["snapshot_player_count"] = len(snap.get("players", []))
    _write("meta.json", meta)


def export_leaderboard(wb, owner_map, api_lookup):
    """Export leaderboard with milestone-adjusted totals."""
    milestone_totals = compute_owner_totals_with_milestones(
        list(api_lookup.values()), owner_map
    )
    effective_map = apply_replacements_to_owner_map(owner_map)

    # Get last-update label from leaderboard sheet
    ws_lb = wb[SHEET_LEADERBOARD]
    rows = []
    for row in range(2, ws_lb.max_row + 1):
        owner = ws_lb.cell(row=row, column=2).value
        last_pts = ws_lb.cell(row=row, column=4).value
        if owner:
            total = milestone_totals.get(owner, 0)
            rows.append({
                "rank": int(ws_lb.cell(row=row, column=1).value or 0),
                "owner": owner,
                "total_points": round(total, 1),
                "last_update_points": round(last_pts or 0, 1),
                "player_count": len(effective_map.get(owner, [])),
            })

    rows.sort(key=lambda x: x["total_points"], reverse=True)
    for i, r in enumerate(rows, 1):
        r["rank"] = i

    _write("leaderboard.json", {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "owners": rows,
    })


def export_owner_rosters(owner_map, api_lookup):
    """Export each owner's full roster with status and points."""
    draft = load_replacement_draft()
    effective_map = apply_replacements_to_owner_map(owner_map)

    owners_data = {}
    for owner in owner_map:
        current_roster = list(effective_map.get(owner, []))

        # Also include dropped players
        dropped_players = []
        for m in draft.get("milestones", []):
            changes = m.get("replacements", {}).get(owner, {})
            for d in changes.get("dropped", []):
                if d not in changes.get("picked", []) and d not in current_roster:
                    dropped_players.append(d)

        full_roster = current_roster + dropped_players

        players = []
        for p_name in full_roster:
            api_p = api_lookup.get(p_name, {})
            curr_pts = api_p.get("OverallPoints", 0)

            # Determine status and milestone-adjusted points
            splits = compute_milestone_adjusted_points(p_name, curr_pts, owner_map)
            owner_pts = sum(s["points"] for s in splits if s["owner"] == owner)
            phases = [s for s in splits if s["owner"] == owner]
            phase_label = phases[0]["phase"] if phases else "—"

            # Status label
            is_dropped = p_name in dropped_players
            is_new = p_name not in owner_map.get(owner, [])
            if is_dropped:
                status = "dropped"
            elif is_new:
                status = "replacement"
            else:
                status = "original"

            players.append({
                "name": p_name,
                "team": api_p.get("TeamShortName", "—"),
                "role": api_p.get("SkillName", "—"),
                "credits": api_p.get("Value", 0),
                "total_api_points": round(curr_pts, 1),
                "owner_points": round(owner_pts, 1),
                "status": status,
                "phase": phase_label,
            })

        players.sort(key=lambda x: x["owner_points"], reverse=True)
        owners_data[owner] = {
            "active_count": len(current_roster),
            "total_points": round(sum(p["owner_points"] for p in players), 1),
            "players": players,
        }

    _write("owner_rosters.json", owners_data)


def export_all_players(api_lookup, owner_map):
    """Export every API player with ownership info."""
    draft = load_replacement_draft()
    effective_map = apply_replacements_to_owner_map(owner_map)
    effective_p2o = build_player_to_owner(effective_map)
    original_p2o = build_player_to_owner(owner_map)

    players = []
    for p in api_lookup.values():
        name = p.get("Name", "")
        curr_pts = p.get("OverallPoints", 0)
        current_owner = effective_p2o.get(name)
        original_owner = original_p2o.get(name)

        # Milestone-adjusted points if owned
        owner_pts = None
        if current_owner:
            splits = compute_milestone_adjusted_points(name, curr_pts, owner_map)
            owner_pts = round(sum(s["points"] for s in splits if s["owner"] == current_owner), 1)

        # Ownership status
        if current_owner and original_owner == current_owner:
            ownership_status = "original"
        elif current_owner and original_owner != current_owner:
            ownership_status = "replacement"
        elif not current_owner and original_owner:
            ownership_status = "dropped"
        else:
            ownership_status = "unowned"

        players.append({
            "name": name,
            "team": p.get("TeamShortName", "—"),
            "role": p.get("SkillName", "—"),
            "credits": p.get("Value", 0),
            "total_points": round(curr_pts, 1),
            "owner": current_owner,
            "original_owner": original_owner,
            "owner_points": owner_pts,
            "ownership_status": ownership_status,
        })

    players.sort(key=lambda x: x["total_points"], reverse=True)
    _write("all_players.json", {
        "total": len(players),
        "players": players,
    })


def export_fixtures(wb):
    """Export fixture tracker data."""
    ws = wb[SHEET_FIXTURES]
    fixtures = []
    for row in range(2, ws.max_row + 1):
        match_num = ws.cell(row=row, column=1).value
        if not match_num:
            continue
        fixtures.append({
            "match": match_num,
            "gameday_id": ws.cell(row=row, column=2).value,
            "date": str(ws.cell(row=row, column=3).value or ""),
            "home": ws.cell(row=row, column=4).value,
            "away": ws.cell(row=row, column=5).value,
            "venue": ws.cell(row=row, column=6).value,
            "result": ws.cell(row=row, column=7).value,
            "status": ws.cell(row=row, column=8).value,
        })
    _write("fixtures.json", {"fixtures": fixtures})


def export_replacement_config():
    """Export the replacement draft config for the FE config editor."""
    draft = load_replacement_draft()
    owner_map = load_owner_map()

    # Enrich with snapshot status
    for m in draft.get("milestones", []):
        snap = load_milestone_snapshot(m["after_match"])
        m["snapshot_saved"] = snap is not None
        m["snapshot_player_count"] = len(snap) if snap else 0

    _write("replacement_config.json", {
        "draft": draft,
        "owner_map": owner_map,
    })


def export_match_history(wb, owner_map):
    """Export per-owner match-by-match point history for charts."""
    history = {}
    for owner in owner_map:
        sheet_name = f"{OWNER_SHEET_PREFIX}{owner}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Get all column labels (row 1, col 5 onwards = match updates)
        labels = []
        for col in range(6, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                labels.append((col, str(val)))

        match_totals = {}
        for col, label in labels:
            col_total = 0
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if isinstance(val, (int, float)):
                    col_total += val
            match_totals[label] = round(col_total, 1)

        history[owner] = match_totals

    _write("match_history.json", history)


def export_config_sources():
    """Copy the raw JSON config files so the JS scoring engine can use them."""
    # 1. Base config files
    for src in [OWNER_MAP_JSON, REPLACEMENT_DRAFT_JSON]:
        if os.path.exists(src):
            dst = os.path.join(WEB_DATA_DIR, os.path.basename(src))
            shutil.copy2(src, dst)
            print(f"   📂 Copied {os.path.basename(src)}")

    # 2. Milestone snapshots
    snap_dir = os.path.join(WEB_DATA_DIR, "milestone_snapshots")
    os.makedirs(snap_dir, exist_ok=True)
    if os.path.exists(MILESTONE_SNAPSHOTS_DIR):
        for f in os.listdir(MILESTONE_SNAPSHOTS_DIR):
            if f.endswith(".json"):
                shutil.copy2(os.path.join(MILESTONE_SNAPSHOTS_DIR, f), os.path.join(snap_dir, f))
                print(f"   📂 Copied snapshot: {f}")


def run_export():
    print("\n🌐 Exporting data for web app...")
    if not os.path.exists(LEDGER_FILE):
        print("❌ Ledger not found. Run 'python main.py --run' first.")
        return False

    owner_map = load_owner_map()
    wb = openpyxl.load_workbook(LEDGER_FILE, data_only=True)

    # Build API lookup from snapshot
    api_lookup = {}
    if os.path.exists(API_SNAPSHOT):
        with open(API_SNAPSHOT) as f:
            snap = json.load(f)
        api_lookup = {p["Name"]: p for p in snap.get("players", [])}

    export_meta()
    export_leaderboard(wb, owner_map, api_lookup)
    export_owner_rosters(owner_map, api_lookup)
    export_all_players(api_lookup, owner_map)
    export_fixtures(wb)
    export_replacement_config()
    export_match_history(wb, owner_map)
    export_config_sources()

    print(f"✅ Web data exported to web/data/")
    return True


if __name__ == "__main__":
    run_export()
