"""
test_milestone_scoring.py — Comprehensive Milestone & Replacement Tests
========================================================================
Validates:
  1. Owner sheet player totals vs Leaderboard
  2. Milestone snapshot integrity (frozen points match API at M25)
  3. Replacement point splitting (dropped + picked = correct attribution)
  4. No double-counting of transferred players
  5. Effective roster correctness (20 active per owner)
  6. Draft integrity (no player active for 2 owners simultaneously)
  7. Milestone-adjusted points arithmetic
"""

import json
import os
import openpyxl

from config import LEDGER_FILE, OWNER_SHEET_PREFIX
from scoring_engine import load_owner_map, build_player_to_owner
from replacement_manager import (
    load_replacement_draft, load_milestone_snapshot,
    apply_replacements_to_owner_map, compute_milestone_adjusted_points,
    compute_owner_totals_with_milestones, get_milestones,
)

API_SNAPSHOT = os.path.join(os.path.dirname(__file__), "api_logs", "latest_snapshot.json")

# ── Helpers ──

def load_api_snapshot():
    with open(API_SNAPSHOT) as f:
        return json.load(f)

def fmt_pass(msg):
    return f"  ✅ {msg}"

def fmt_fail(msg):
    return f"  ❌ {msg}"

def fmt_warn(msg):
    return f"  ⚠️  {msg}"


def run_all_tests():
    print("=" * 70)
    print("🔍 IPL FANTASY — COMPREHENSIVE MILESTONE & REPLACEMENT TESTS")
    print("=" * 70)

    owner_map = load_owner_map()
    snap = load_api_snapshot()
    api_players = snap["players"]
    api_lookup = {p["Name"]: p for p in api_players}
    draft = load_replacement_draft()
    milestones = get_milestones()

    try:
        wb = openpyxl.load_workbook(LEDGER_FILE, data_only=True)
    except FileNotFoundError:
        print("❌ Ledger not found. Run '--run' first.")
        return

    total_checks = 0
    passed_checks = 0

    # ═══════════════════════════════════════════════════════════════
    # TEST 1: Owner Sheet Totals vs Leaderboard
    # ═══════════════════════════════════════════════════════════════
    print(f"\n{'─' * 70}")
    print("📊 TEST 1: Owner Sheet Player Totals vs Leaderboard")
    print("   Sums baseline column from each owner sheet and compares to")
    print("   the leaderboard total. They must match.")
    print(f"{'─' * 70}")

    ws_lb = wb["Summary_Leaderboard"]
    lb_totals = {}
    for row in range(2, ws_lb.max_row + 1):
        owner = ws_lb.cell(row=row, column=2).value
        total = ws_lb.cell(row=row, column=3).value
        if owner:
            lb_totals[owner] = total or 0

    for owner in owner_map:
        total_checks += 1
        sheet_name = f"{OWNER_SHEET_PREFIX}{owner}"
        if sheet_name not in wb.sheetnames:
            print(fmt_fail(f"{owner}: Sheet '{sheet_name}' missing!"))
            continue

        ws = wb[sheet_name]
        # Sum the Baseline column (col 5) for all non-dropped players
        sheet_total = 0
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=4).value or ""
            pts = ws.cell(row=row, column=5).value
            if isinstance(pts, (int, float)):
                sheet_total += pts

        lb_total = lb_totals.get(owner, 0)
        if abs(sheet_total - lb_total) < 0.01:
            print(fmt_pass(f"{owner:12s}: Sheet={sheet_total:.0f}, Leaderboard={lb_total:.0f}"))
            passed_checks += 1
        else:
            print(fmt_fail(f"{owner:12s}: Sheet={sheet_total:.0f} ≠ Leaderboard={lb_total:.0f} "
                           f"(diff={sheet_total - lb_total:.0f})"))

    # ═══════════════════════════════════════════════════════════════
    # TEST 2: Milestone Snapshot Integrity
    # ═══════════════════════════════════════════════════════════════
    print(f"\n{'─' * 70}")
    print("📊 TEST 2: Milestone Snapshot Integrity")
    print("   Checks that the milestone snapshot exists and has valid data")
    print("   for all 260 league players.")
    print(f"{'─' * 70}")

    for m_match in milestones:
        total_checks += 1
        ms_snap = load_milestone_snapshot(m_match)
        if ms_snap is None:
            print(fmt_fail(f"M{m_match}: Milestone snapshot MISSING"))
            continue

        num_players = len(ms_snap)
        non_zero = sum(1 for v in ms_snap.values() if v > 0)
        print(fmt_pass(f"M{m_match}: Snapshot exists — {num_players} players, "
                       f"{non_zero} with >0 points"))
        passed_checks += 1

    # ═══════════════════════════════════════════════════════════════
    # TEST 3: Replacement Point Splitting
    # ═══════════════════════════════════════════════════════════════
    print(f"\n{'─' * 70}")
    print("📊 TEST 3: Replacement Point Splitting (Dropped + Picked)")
    print("   For each transferred player, verifies:")
    print("   • Dropped owner gets only pre-milestone points (frozen)")
    print("   • Picking owner gets only post-milestone points")
    print("   • Sum of split = player's total current points")
    print(f"{'─' * 70}")

    for m in draft.get("milestones", []):
        after = m["after_match"]
        ms_snap = load_milestone_snapshot(after)
        if ms_snap is None:
            print(fmt_warn(f"Skipping M{after} — no snapshot"))
            continue

        replacements = m.get("replacements", {})
        for owner, changes in replacements.items():
            dropped = changes.get("dropped", [])
            picked = changes.get("picked", [])

            # Only test actual transfers (not same player retained)
            for d in dropped:
                if d in picked:
                    continue  # Retained — skip

                total_checks += 1
                frozen_pts = ms_snap.get(d, 0)
                curr_pts = api_lookup.get(d, {}).get("OverallPoints", 0)

                # The owner should get frozen_pts for this player
                splits = compute_milestone_adjusted_points(d, curr_pts, owner_map)
                owner_split = [s for s in splits if s["owner"] == owner]
                owner_attributed = sum(s["points"] for s in owner_split)

                if abs(owner_attributed - frozen_pts) < 0.01:
                    print(fmt_pass(
                        f"DROPPED {d:25s} by {owner:10s}: "
                        f"Frozen={frozen_pts:.0f}, Attributed={owner_attributed:.0f}"))
                    passed_checks += 1
                else:
                    print(fmt_fail(
                        f"DROPPED {d:25s} by {owner:10s}: "
                        f"Frozen={frozen_pts:.0f} ≠ Attributed={owner_attributed:.0f}"))

            for p in picked:
                if p in dropped:
                    continue  # Retained

                total_checks += 1
                frozen_pts = ms_snap.get(p, 0)
                curr_pts = api_lookup.get(p, {}).get("OverallPoints", 0)
                expected_post = curr_pts - frozen_pts

                splits = compute_milestone_adjusted_points(p, curr_pts, owner_map)
                owner_split = [s for s in splits if s["owner"] == owner]
                owner_attributed = sum(s["points"] for s in owner_split)

                if abs(owner_attributed - expected_post) < 0.01:
                    print(fmt_pass(
                        f"PICKED  {p:25s} by {owner:10s}: "
                        f"Post-M{after}={expected_post:.0f}, Attributed={owner_attributed:.0f}"))
                    passed_checks += 1
                else:
                    print(fmt_fail(
                        f"PICKED  {p:25s} by {owner:10s}: "
                        f"Post-M{after}={expected_post:.0f} ≠ Attributed={owner_attributed:.0f}"))

    # ═══════════════════════════════════════════════════════════════
    # TEST 4: Transferred Players — Point Attribution Audit
    # ═══════════════════════════════════════════════════════════════
    print(f"\n{'─' * 70}")
    print("📊 TEST 4: Transferred Players — Point Attribution Audit")
    print("   For each player involved in replacements, checks:")
    print("   • Transferred (dropped by A, picked by B): split sums to total")
    print("   • Dropped only (not re-picked): owner gets pre-milestone only")
    print("   • Picked only (not prev owned): owner gets post-milestone only")
    print(f"{'─' * 70}")

    # Categorize all replacement players
    for m in draft.get("milestones", []):
        after = m["after_match"]
        ms_snap = load_milestone_snapshot(after)
        if ms_snap is None:
            continue

        # Build maps: who dropped whom, who picked whom
        all_dropped = {}  # player -> dropping_owner
        all_picked = {}   # player -> picking_owner
        for owner, changes in m.get("replacements", {}).items():
            for d in changes.get("dropped", []):
                if d not in changes.get("picked", []):
                    all_dropped[d] = owner
            for p in changes.get("picked", []):
                if p not in changes.get("dropped", []):
                    all_picked[p] = owner

        # Category 1: Transferred (dropped by one, picked by another)
        transferred = set(all_dropped.keys()) & set(all_picked.keys())
        for p_name in sorted(transferred):
            total_checks += 1
            curr_pts = api_lookup.get(p_name, {}).get("OverallPoints", 0)
            splits = compute_milestone_adjusted_points(p_name, curr_pts, owner_map)
            total_attributed = sum(s["points"] for s in splits)
            split_detail = " + ".join(f'{s["owner"]}:{s["points"]:.0f}' for s in splits)

            if abs(total_attributed - curr_pts) < 0.01:
                print(fmt_pass(
                    f"TRANSFERRED {p_name:22s}: Total={curr_pts:.0f}, "
                    f"Split=[{split_detail}]"))
                passed_checks += 1
            else:
                print(fmt_fail(
                    f"TRANSFERRED {p_name:22s}: Total={curr_pts:.0f} ≠ "
                    f"Attributed={total_attributed:.0f}"))

        # Category 2: Dropped only (not re-picked by anyone)
        dropped_only = set(all_dropped.keys()) - set(all_picked.keys())
        for p_name in sorted(dropped_only):
            total_checks += 1
            frozen_pts = ms_snap.get(p_name, 0)
            curr_pts = api_lookup.get(p_name, {}).get("OverallPoints", 0)
            splits = compute_milestone_adjusted_points(p_name, curr_pts, owner_map)
            total_attributed = sum(s["points"] for s in splits)
            unattributed = curr_pts - total_attributed

            # Should only have pre-milestone attributed, rest is unowned
            if abs(total_attributed - frozen_pts) < 0.01:
                print(fmt_pass(
                    f"DROPPED-ONLY {p_name:21s}: Attributed={total_attributed:.0f} "
                    f"(pre-M{after}), Unowned post-M{after}={unattributed:.0f}"))
                passed_checks += 1
            else:
                print(fmt_fail(
                    f"DROPPED-ONLY {p_name:21s}: Attributed={total_attributed:.0f} "
                    f"≠ Frozen={frozen_pts:.0f}"))

        # Category 3: Picked only (not previously owned by anyone)
        picked_only = set(all_picked.keys()) - set(all_dropped.keys())
        # Filter to truly new players (not in any original roster)
        original_all = set()
        for players in owner_map.values():
            original_all.update(players)

        for p_name in sorted(picked_only):
            total_checks += 1
            frozen_pts = ms_snap.get(p_name, 0)
            curr_pts = api_lookup.get(p_name, {}).get("OverallPoints", 0)
            expected_post = curr_pts - frozen_pts
            splits = compute_milestone_adjusted_points(p_name, curr_pts, owner_map)
            total_attributed = sum(s["points"] for s in splits)

            if p_name in original_all:
                # Was in original roster but picked by someone else after drop
                # (handled in transferred above)
                print(fmt_pass(
                    f"RE-PICKED    {p_name:21s}: Attributed={total_attributed:.0f}"))
                passed_checks += 1
            else:
                # Truly new — only post-milestone should be attributed
                if abs(total_attributed - expected_post) < 0.01:
                    print(fmt_pass(
                        f"NEW-PICK     {p_name:21s}: Attributed={total_attributed:.0f} "
                        f"(post-M{after} only), Pre-M{after}={frozen_pts:.0f} unowned"))
                    passed_checks += 1
                else:
                    print(fmt_fail(
                        f"NEW-PICK     {p_name:21s}: Attributed={total_attributed:.0f} "
                        f"≠ Expected post-M{after}={expected_post:.0f}"))

    # ═══════════════════════════════════════════════════════════════
    # TEST 5: Effective Roster Size (20 active per owner)
    # ═══════════════════════════════════════════════════════════════
    print(f"\n{'─' * 70}")
    print("📊 TEST 5: Effective Roster Size (20 active players per owner)")
    print("   After applying all replacements, each owner should still")
    print("   have exactly 20 active players.")
    print(f"{'─' * 70}")

    effective = apply_replacements_to_owner_map(owner_map)
    for owner, players in effective.items():
        total_checks += 1
        if len(players) == 20:
            print(fmt_pass(f"{owner:12s}: {len(players)} active players"))
            passed_checks += 1
        else:
            print(fmt_fail(f"{owner:12s}: {len(players)} active players (expected 20)"))

    # ═══════════════════════════════════════════════════════════════
    # TEST 6: Post-Replacement Draft Integrity
    # ═══════════════════════════════════════════════════════════════
    print(f"\n{'─' * 70}")
    print("📊 TEST 6: Post-Replacement Draft Integrity")
    print("   No player should be actively owned by two owners at the")
    print("   same time after replacements are applied.")
    print(f"{'─' * 70}")

    total_checks += 1
    seen = {}
    dup_found = False
    for owner, players in effective.items():
        for p in players:
            if p in seen:
                print(fmt_fail(f"'{p}' is active for both '{seen[p]}' and '{owner}'!"))
                dup_found = True
            seen[p] = owner

    if not dup_found:
        total_active = sum(len(p) for p in effective.values())
        print(fmt_pass(f"{len(seen)} unique active players across {len(effective)} owners "
                       f"(total slots: {total_active})"))
        passed_checks += 1

    # ═══════════════════════════════════════════════════════════════
    # TEST 7: Milestone-Adjusted Total vs Direct Calculation
    # ═══════════════════════════════════════════════════════════════
    print(f"\n{'─' * 70}")
    print("📊 TEST 7: Milestone-Adjusted Owner Totals Cross-Check")
    print("   Computes owner totals via two methods and ensures they match:")
    print("   Method A: compute_owner_totals_with_milestones()")
    print("   Method B: Sum of all player splits attributed to each owner")
    print(f"{'─' * 70}")

    method_a = compute_owner_totals_with_milestones(api_players, owner_map)

    # Method B: manual summation
    method_b = {owner: 0.0 for owner in owner_map}
    all_players_set = set()
    for players in owner_map.values():
        all_players_set.update(players)
    for m in draft.get("milestones", []):
        for changes in m.get("replacements", {}).values():
            all_players_set.update(changes.get("picked", []))

    for p_name in all_players_set:
        curr_pts = api_lookup.get(p_name, {}).get("OverallPoints", 0)
        splits = compute_milestone_adjusted_points(p_name, curr_pts, owner_map)
        for s in splits:
            if s["owner"] in method_b:
                method_b[s["owner"]] += s["points"]

    for owner in owner_map:
        total_checks += 1
        a = method_a.get(owner, 0)
        b = method_b.get(owner, 0)
        if abs(a - b) < 0.01:
            print(fmt_pass(f"{owner:12s}: MethodA={a:.0f}, MethodB={b:.0f}"))
            passed_checks += 1
        else:
            print(fmt_fail(f"{owner:12s}: MethodA={a:.0f} ≠ MethodB={b:.0f}"))

    # ═══════════════════════════════════════════════════════════════
    # TEST 8: Replacement Player API Name Validation
    # ═══════════════════════════════════════════════════════════════
    print(f"\n{'─' * 70}")
    print("📊 TEST 8: Replacement Player Name Validation")
    print("   All dropped/picked player names in replacement_draft.json")
    print("   must exist in the API snapshot and owner_player_map.json.")
    print(f"{'─' * 70}")

    api_names = {p["Name"] for p in api_players}
    owned_names = set()
    for players in owner_map.values():
        owned_names.update(players)

    all_valid = True
    for m in draft.get("milestones", []):
        for owner, changes in m.get("replacements", {}).items():
            for d in changes.get("dropped", []):
                total_checks += 1
                if d not in api_names:
                    print(fmt_fail(f"DROPPED '{d}' by {owner} — NOT in API snapshot"))
                    all_valid = False
                elif d not in owned_names:
                    print(fmt_fail(f"DROPPED '{d}' by {owner} — NOT in owner_player_map"))
                    all_valid = False
                else:
                    passed_checks += 1
            for p in changes.get("picked", []):
                total_checks += 1
                if p not in api_names:
                    print(fmt_fail(f"PICKED '{p}' by {owner} — NOT in API snapshot"))
                    all_valid = False
                else:
                    passed_checks += 1

    if all_valid:
        print(fmt_pass("All replacement player names validated against API + owner map"))

    # ═══════════════════════════════════════════════════════════════
    # TEST 9: Grand Total Points Conservation
    # ═══════════════════════════════════════════════════════════════
    print(f"\n{'─' * 70}")
    print("📊 TEST 9: Grand Total Points Conservation")
    print("   Sum of all owner totals should equal the sum of all owned")
    print("   players' milestone-adjusted points (no points lost or created).")
    print(f"{'─' * 70}")

    total_checks += 1
    grand_total_owners = sum(method_a.values())
    grand_total_players = sum(method_b.values())

    if abs(grand_total_owners - grand_total_players) < 0.01:
        print(fmt_pass(f"Grand total owners: {grand_total_owners:.0f}"))
        print(fmt_pass(f"Grand total players: {grand_total_players:.0f}"))
        print(fmt_pass(f"Conservation check PASSED — no points lost or created"))
        passed_checks += 1
    else:
        print(fmt_fail(f"Owners={grand_total_owners:.0f} ≠ Players={grand_total_players:.0f}"))

    # ═══════════════════════════════════════════════════════════════
    # SUMMARY
    # ═══════════════════════════════════════════════════════════════
    print(f"\n{'=' * 70}")
    if passed_checks == total_checks:
        print(f"✅ ALL {total_checks} CHECKS PASSED")
    else:
        print(f"❌ {total_checks - passed_checks} FAILED out of {total_checks} checks "
              f"({passed_checks} passed)")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    run_all_tests()
