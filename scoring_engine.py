"""
scoring_engine.py — Hybrid Scoring Logic
=========================================
Handles two modes:
  - Bulk (≤ cutoff): Ingest cumulative totals from the API snapshot.
  - Granular (> cutoff): Track per-match deltas for each player.

Reads/writes to the Excel ledger via openpyxl.
"""

import json
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import (
    OWNER_MAP_JSON, LEDGER_FILE, BULK_CUTOFF_MATCH,
    SHEET_LEADERBOARD, SHEET_GLOBAL_LEDGER, SHEET_FIXTURES,
    OWNER_SHEET_PREFIX,
)
from replacement_manager import (
    load_replacement_draft, get_milestones, apply_replacements_to_owner_map,
    compute_milestone_adjusted_points, compute_owner_totals_with_milestones,
    load_milestone_snapshot,
)

SHEET_ALL_PLAYERS = "All_Owned_Players"


# ── Errors ──
class DraftIntegrityError(Exception):
    """Raised when a player is mapped to multiple owners."""
    pass


# ── Styles ──
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
OWNER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
RANK1_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center")


def _style_header(ws, headers, row=1):
    """Apply header styling to a row."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def load_owner_map() -> dict:
    """
    Load owner→players mapping from JSON.
    Validates no player appears under multiple owners.
    Returns: {owner_name: [player_names]}
    """
    with open(OWNER_MAP_JSON, "r") as f:
        mapping = json.load(f)

    # Integrity check
    seen = {}
    for owner, players in mapping.items():
        for p in players:
            if p in seen:
                raise DraftIntegrityError(
                    f"Player '{p}' is mapped to both '{seen[p]}' and '{owner}'!"
                )
            seen[p] = owner

    return mapping


def build_player_to_owner(owner_map: dict) -> dict:
    """Invert the owner map: {player_name: owner_name}."""
    return {p: owner for owner, players in owner_map.items() for p in players}


def init_ledger():
    """
    Create a fresh ledger workbook with all required sheets.
    If the file exists, load it instead.
    """
    if os.path.exists(LEDGER_FILE):
        return openpyxl.load_workbook(LEDGER_FILE)

    wb = openpyxl.Workbook()

    # ── 1. Summary_Leaderboard ──
    ws = wb.active
    ws.title = SHEET_LEADERBOARD
    _style_header(ws, ["Rank", "Owner", "Total Points", "Last Match Points", "# Players"])
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A2"

    # ── 2. Global_Match_Ledger ──
    ws_gl = wb.create_sheet(SHEET_GLOBAL_LEDGER)
    _style_header(ws_gl, [
        "Update", "Timestamp", "Player Name", "Owner",
        "Team", "Role", "Delta Points", "Cumulative Points",
    ])
    for col_letter, width in [("A", 12), ("B", 18), ("C", 26), ("D", 12),
                               ("E", 6), ("F", 14), ("G", 13), ("H", 16)]:
        ws_gl.column_dimensions[col_letter].width = width
    ws_gl.freeze_panes = "A2"

    # ── 3. Fixture_Tracker ──
    ws_ft = wb.create_sheet(SHEET_FIXTURES)
    _style_header(ws_ft, [
        "Match #", "Gameday ID", "Date", "Home", "Away",
        "Venue", "Result", "Status",
    ])
    for col_letter, width in [("A", 9), ("B", 12), ("C", 12), ("D", 10),
                               ("E", 10), ("F", 30), ("G", 40), ("H", 12)]:
        ws_ft.column_dimensions[col_letter].width = width
    ws_ft.freeze_panes = "A2"

    # ── 4. Owner sheets ──
    owner_map = load_owner_map()
    for owner in owner_map:
        ws_owner = wb.create_sheet(f"{OWNER_SHEET_PREFIX}{owner}")
        headers = ["Player Name", "Team", "Role", "Status", "Baseline"]
        _style_header(ws_owner, headers)
        ws_owner.column_dimensions["A"].width = 26
        ws_owner.column_dimensions["B"].width = 6
        ws_owner.column_dimensions["C"].width = 14
        ws_owner.column_dimensions["D"].width = 20
        ws_owner.column_dimensions["E"].width = 16
        ws_owner.freeze_panes = "A2"

    wb.save(LEDGER_FILE)
    print(f"📒 Created new ledger: {LEDGER_FILE}")
    return wb


def _get_player_status_and_points(player_name, owner, owner_map, api_lookup):
    """
    Determine a player's status label and milestone-adjusted points.
    Returns (status_label, adjusted_points).
    """
    milestones = get_milestones()
    if not milestones:
        pts = api_lookup.get(player_name, {}).get("OverallPoints", 0)
        return ("Original", pts)

    draft = load_replacement_draft()
    original_players = set(owner_map.get(owner, []))

    # Check if player was dropped or picked in any milestone
    for m in draft.get("milestones", []):
        changes = m.get("replacements", {}).get(owner, {})
        dropped = changes.get("dropped", [])
        picked = changes.get("picked", [])

        if player_name in picked and player_name not in original_players:
            # New pick — only post-milestone points
            milestone_snap = load_milestone_snapshot(m["after_match"])
            if milestone_snap:
                frozen_pts = milestone_snap.get(player_name, 0)
                curr_pts = api_lookup.get(player_name, {}).get("OverallPoints", 0)
                return (f"Replacement (M{m['after_match']+1}+)", curr_pts - frozen_pts)
            else:
                return (f"Replacement (M{m['after_match']+1}+)", 0)

        if player_name in dropped and player_name not in picked:
            # Dropped — only pre-milestone points
            milestone_snap = load_milestone_snapshot(m["after_match"])
            if milestone_snap:
                frozen_pts = milestone_snap.get(player_name, 0)
                return (f"Dropped (till M{m['after_match']})", frozen_pts)
            else:
                pts = api_lookup.get(player_name, {}).get("OverallPoints", 0)
                return (f"Dropped (till M{m['after_match']})", pts)

    # Original player, not replaced
    pts = api_lookup.get(player_name, {}).get("OverallPoints", 0)
    return ("Original", pts)


def _get_effective_roster(owner, owner_map):
    """
    Get the full roster for display: original players (including dropped ones)
    plus newly picked replacements, in order.
    """
    effective = apply_replacements_to_owner_map(owner_map)
    current_roster = effective.get(owner, [])

    # Also include dropped players for historical visibility
    draft = load_replacement_draft()
    dropped_players = []
    for m in draft.get("milestones", []):
        changes = m.get("replacements", {}).get(owner, {})
        for d in changes.get("dropped", []):
            picked = changes.get("picked", [])
            if d not in picked and d not in current_roster:
                dropped_players.append(d)

    return current_roster + dropped_players


# ── Styling for status labels ──
DROPPED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
REPLACEMENT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def populate_fixtures(wb: openpyxl.Workbook, fixtures: list):
    """Write all 70 fixtures to the Fixture_Tracker sheet."""
    ws = wb[SHEET_FIXTURES]

    # Clear existing data (keep header)
    for row in range(2, ws.max_row + 1):
        for col in range(1, 9):
            ws.cell(row=row, column=col).value = None

    for i, fix in enumerate(fixtures, 2):
        match_num = fix.get("Gameday", "")
        status = "Completed" if fix.get("MatchStatus") == 2 else "Pending"
        values = [
            match_num,
            fix.get("TourGamedayId", ""),
            fix.get("Matchdate", ""),
            fix.get("HomeTeamShortName", ""),
            fix.get("AwayTeamShortName", ""),
            fix.get("Venue", ""),
            fix.get("MatchResult", "") if status == "Completed" else "",
            status,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER
            if col == 8:
                cell.fill = PatternFill(
                    start_color="C6EFCE" if status == "Completed" else "FFEB9C",
                    end_color="C6EFCE" if status == "Completed" else "FFEB9C",
                    fill_type="solid",
                )

    wb.save(LEDGER_FILE)


def _write_owner_sheet_rows(ws, owner, owner_map, api_lookup, data_col_start=5):
    """
    Write player rows to an owner sheet with status labels.
    Handles original, replacement, and dropped players.
    data_col_start is the first column for point data (after Name/Team/Role/Status).
    """
    roster = _get_effective_roster(owner, owner_map)

    # Clear data rows
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None

    for i, player_name in enumerate(roster, 2):
        api_p = api_lookup.get(player_name, {})
        status_label, adjusted_pts = _get_player_status_and_points(
            player_name, owner, owner_map, api_lookup
        )

        ws.cell(row=i, column=1, value=player_name).border = THIN_BORDER
        ws.cell(row=i, column=2, value=api_p.get("TeamShortName", "-")).border = THIN_BORDER
        ws.cell(row=i, column=3, value=api_p.get("SkillName", "-")).border = THIN_BORDER
        ws.cell(row=i, column=4, value=status_label).border = THIN_BORDER

        # Color the status cell
        if "Dropped" in status_label:
            ws.cell(row=i, column=4).fill = DROPPED_FILL
            ws.cell(row=i, column=4).font = Font(color="9C0006")
        elif "Replacement" in status_label:
            ws.cell(row=i, column=4).fill = REPLACEMENT_FILL
            ws.cell(row=i, column=4).font = Font(color="006100")

        ws.cell(row=i, column=data_col_start, value=adjusted_pts).border = THIN_BORDER

    return roster


def ingest_bulk(wb: openpyxl.Workbook, api_players: list, owner_map: dict):
    """
    Bulk mode: Write milestone-adjusted points for each owned player.
    Shows effective roster with status labels.
    """
    player_to_owner = build_player_to_owner(owner_map)
    api_lookup = {p["Name"]: p for p in api_players}

    # ── Update Owner Sheets with effective roster ──
    for owner in owner_map:
        sheet_name = f"{OWNER_SHEET_PREFIX}{owner}"
        ws = wb[sheet_name]

        # Ensure header has Status + Baseline columns
        headers = ["Player Name", "Team", "Role", "Status", "Baseline"]
        _style_header(ws, headers)
        ws.column_dimensions["D"].width = 20

        _write_owner_sheet_rows(ws, owner, owner_map, api_lookup, data_col_start=5)

    # ── Write to Global Ledger as bulk entry ──
    ws_gl = wb[SHEET_GLOBAL_LEDGER]
    for row in range(2, ws_gl.max_row + 1):
        for col in range(1, 9):
            ws_gl.cell(row=row, column=col).value = None

    row = 2
    for player_name, owner in sorted(player_to_owner.items(), key=lambda x: x[1]):
        api_p = api_lookup.get(player_name, {})
        pts = api_p.get("OverallPoints", 0)
        if pts == 0:
            continue
        values = [
            f"≤M{BULK_CUTOFF_MATCH}",
            "",
            player_name,
            owner,
            api_p.get("TeamShortName", "-"),
            api_p.get("SkillName", "-"),
            pts,
            pts,
        ]
        for col, val in enumerate(values, 1):
            ws_gl.cell(row=row, column=col, value=val).border = THIN_BORDER
        row += 1

    # ── Update Leaderboard ──
    _update_leaderboard(wb, owner_map, api_lookup)

    # ── Update All Owned Players sheet ──
    update_all_players_sheet(wb, owner_map, api_lookup)

    wb.save(LEDGER_FILE)
    print(f"✅ Baseline ingestion complete")


def ingest_update(wb: openpyxl.Workbook, label: str, api_players: list,
                  prev_players: list, owner_map: dict, fixtures: list):
    """
    Compute per-player deltas between current and previous snapshot.
    Uses effective roster (with replacements) and status labels.
    Label is a date string like '14-Apr'.
    """
    from datetime import datetime

    curr_lookup = {p["Name"]: p for p in api_players}
    prev_lookup = {p["Name"]: p for p in prev_players}
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Rebuild Owner Sheets with effective roster, then add delta column ──
    for owner in owner_map:
        sheet_name = f"{OWNER_SHEET_PREFIX}{owner}"
        ws = wb[sheet_name]

        # Rebuild rows with status labels and baseline
        headers = ["Player Name", "Team", "Role", "Status", "Baseline"]
        _style_header(ws, headers)
        ws.column_dimensions["D"].width = 20
        roster = _write_owner_sheet_rows(ws, owner, owner_map, curr_lookup, data_col_start=5)

        # Find next available column for the delta label (or reuse)
        next_col = ws.max_column + 1
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == label:
                next_col = col
                break
        else:
            cell = ws.cell(row=1, column=next_col, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            ws.column_dimensions[openpyxl.utils.get_column_letter(next_col)].width = 12

        for i, player_name in enumerate(roster, 2):
            status_label = ws.cell(row=i, column=4).value or ""
            if "Dropped" in status_label:
                # Dropped players get 0 delta — they're no longer scoring
                delta = 0
            else:
                curr_pts = curr_lookup.get(player_name, {}).get("OverallPoints", 0)
                prev_pts = prev_lookup.get(player_name, {}).get("OverallPoints", 0)
                delta = curr_pts - prev_pts
            ws.cell(row=i, column=next_col, value=delta).border = THIN_BORDER

    # ── Append to Global Ledger ──
    # Build effective player→owner mapping for the global ledger
    effective_map = apply_replacements_to_owner_map(owner_map)
    effective_p2o = build_player_to_owner(effective_map)

    ws_gl = wb[SHEET_GLOBAL_LEDGER]
    next_row = ws_gl.max_row + 1

    for player_name, owner in sorted(effective_p2o.items(), key=lambda x: x[1]):
        curr = curr_lookup.get(player_name, {})
        prev = prev_lookup.get(player_name, {})
        curr_pts = curr.get("OverallPoints", 0)
        prev_pts = prev.get("OverallPoints", 0)
        delta = curr_pts - prev_pts

        if delta == 0:
            continue

        values = [
            label,
            timestamp,
            player_name,
            owner,
            curr.get("TeamShortName", "-"),
            curr.get("SkillName", "-"),
            delta,
            curr_pts,
        ]
        for col, val in enumerate(values, 1):
            ws_gl.cell(row=next_row, column=col, value=val).border = THIN_BORDER
        next_row += 1

    # ── Update Leaderboard ──
    _update_leaderboard(wb, owner_map, curr_lookup, label)

    # ── Update All Owned Players sheet ──
    update_all_players_sheet(wb, owner_map, curr_lookup)

    wb.save(LEDGER_FILE)
    print(f"✅ Update '{label}' recorded")


def _update_leaderboard(wb, owner_map, api_lookup, last_label=None):
    """Recalculate and write the Summary_Leaderboard sheet (milestone-aware)."""
    ws = wb[SHEET_LEADERBOARD]

    # Clear data rows
    for row in range(2, ws.max_row + 1):
        for col in range(1, 6):
            ws.cell(row=row, column=col).value = None

    # Calculate totals per owner using milestone-adjusted points
    api_players_list = list(api_lookup.values())
    milestone_totals = compute_owner_totals_with_milestones(api_players_list, owner_map)

    # Get current effective roster size
    effective_map = apply_replacements_to_owner_map(owner_map)

    owner_totals = []
    for owner in owner_map:
        total = milestone_totals.get(owner, 0)

        # Last update points: sum the latest column in the owner sheet
        last_pts = 0
        if last_label:
            sheet_name = f"{OWNER_SHEET_PREFIX}{owner}"
            if sheet_name in wb.sheetnames:
                ws_owner = wb[sheet_name]
                for col in range(1, ws_owner.max_column + 1):
                    if ws_owner.cell(row=1, column=col).value == last_label:
                        for row in range(2, ws_owner.max_row + 1):
                            val = ws_owner.cell(row=row, column=col).value
                            if val and isinstance(val, (int, float)):
                                last_pts += val
                        break

        num_players = len(effective_map.get(owner, []))
        owner_totals.append((owner, total, last_pts, num_players))

    # Sort by total points descending
    owner_totals.sort(key=lambda x: x[1], reverse=True)

    for rank, (owner, total, last_pts, num_players) in enumerate(owner_totals, 1):
        row = rank + 1
        ws.cell(row=row, column=1, value=rank).border = THIN_BORDER
        ws.cell(row=row, column=2, value=owner).border = THIN_BORDER
        ws.cell(row=row, column=3, value=total).border = THIN_BORDER
        ws.cell(row=row, column=4, value=last_pts).border = THIN_BORDER
        ws.cell(row=row, column=5, value=num_players).border = THIN_BORDER

        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=5).alignment = CENTER

        if rank == 1:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = RANK1_FILL


def get_leaderboard(wb: openpyxl.Workbook) -> list:
    """Read the leaderboard from the workbook. Returns list of dicts."""
    ws = wb[SHEET_LEADERBOARD]
    rows = []
    for row in range(2, ws.max_row + 1):
        rank = ws.cell(row=row, column=1).value
        if rank is None:
            break
        rows.append({
            "Rank": rank,
            "Owner": ws.cell(row=row, column=2).value,
            "Total Points": ws.cell(row=row, column=3).value,
            "Last Match Points": ws.cell(row=row, column=4).value,
            "# Players": ws.cell(row=row, column=5).value,
        })
    return rows


def get_player_history(wb: openpyxl.Workbook, player_name: str, owner_map: dict) -> dict:
    """Get a player's match-by-match history from their owner's sheet."""
    player_to_owner = build_player_to_owner(owner_map)
    owner = player_to_owner.get(player_name)
    if not owner:
        return {"error": f"Player '{player_name}' not found in any owner's roster."}

    sheet_name = f"{OWNER_SHEET_PREFIX}{owner}"
    ws = wb[sheet_name]

    # Find the player's row
    player_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == player_name:
            player_row = row
            break

    if player_row is None:
        return {"error": f"Player '{player_name}' not found in {sheet_name}."}

    # Gather match columns
    history = {
        "name": player_name,
        "owner": owner,
        "team": ws.cell(row=player_row, column=2).value,
        "role": ws.cell(row=player_row, column=3).value,
        "matches": {},
    }

    for col in range(4, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        val = ws.cell(row=player_row, column=col).value
        if header and val is not None:
            history["matches"][str(header)] = val

    history["total"] = sum(
        v for v in history["matches"].values() if isinstance(v, (int, float))
    )

    return history


def update_all_players_sheet(wb: openpyxl.Workbook, owner_map: dict, api_lookup: dict):
    """
    Create/update the 'All_Owned_Players' sheet (milestone-aware).
    Lists every owned player with their milestone-adjusted points and current owner.
    """
    # Remove existing sheet if present
    if SHEET_ALL_PLAYERS in wb.sheetnames:
        del wb[SHEET_ALL_PLAYERS]

    # Insert after Summary_Leaderboard
    lb_idx = wb.sheetnames.index(SHEET_LEADERBOARD)
    ws = wb.create_sheet(SHEET_ALL_PLAYERS, lb_idx + 1)

    headers = ["#", "Player Name", "Owner", "Phase", "Team", "Role", "Credits", "Points"]
    _style_header(ws, headers)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 9
    ws.column_dimensions["H"].width = 14
    ws.freeze_panes = "A2"

    # Collect all player-phase entries with milestone-adjusted points
    all_entries = []
    seen_players = set()

    # Gather all players across original + replacements
    for players in owner_map.values():
        seen_players.update(players)
    draft = load_replacement_draft()
    for m in draft.get("milestones", []):
        for changes in m.get("replacements", {}).values():
            seen_players.update(changes.get("picked", []))

    for p_name in seen_players:
        api_p = api_lookup.get(p_name, {})
        current_pts = api_p.get("OverallPoints", 0)
        splits = compute_milestone_adjusted_points(p_name, current_pts, owner_map)
        for split in splits:
            all_entries.append({
                "name": p_name,
                "owner": split["owner"],
                "phase": split["phase"],
                "team": api_p.get("TeamShortName", "-"),
                "role": api_p.get("SkillName", "-"),
                "credits": api_p.get("Value", 0),
                "points": split["points"],
            })

    # Sort by points descending
    all_entries.sort(key=lambda x: x["points"], reverse=True)

    # Write rows
    for i, p in enumerate(all_entries, 2):
        rank = i - 1
        values = [rank, p["name"], p["owner"], p["phase"], p["team"], p["role"], p["credits"], p["points"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER
            if col in (1, 7, 8):
                cell.alignment = CENTER

        # Alternate row shading
        if rank % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = ALT_FILL

        # Gold for top 3
        if rank <= 3:
            for col in range(1, 9):
                ws.cell(row=i, column=col).fill = RANK1_FILL
